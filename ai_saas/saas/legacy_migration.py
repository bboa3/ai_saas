"""Inventory of every tenant site, crossed with what the control site knows.

Phase 1 of the legacy-account migration (decision 2026-08-30): before anything is
registered, archived or campaigned, *look*. The sites are the ground truth — each one
says who it belongs to (Company, its System Managers) and whether it is used — and
the control site's Customers, Contracts, Subscriptions, Opportunities and Leads are
matched **to** them through every key available (site name, NUIT, emails, mobiles,
company name), each match labelled with the key that produced it.

	bench --site <control-site> execute ai_saas.saas.legacy_migration.inventory

writes `tenant_inventory_<YYYYMMDD>.xlsx` (sheets: sites, control_only, summary) as a
private File attached to MZ SaaS Settings — downloadable from the desk — and prints
the URL and the counts. Read-only apart from that File record; idempotent (the same
day's file is replaced).
"""

import datetime
import io
import json
import os
import re
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

import frappe
from frappe.utils import add_days, cint, flt, getdate, nowdate

from ai_saas.saas.provisioning import (
	DOMAINS,
	ProvisioningError,
	get_bench_cmd,
	get_bench_path,
	run_cmd_capture,
)

IDENTITY_METHOD = "erpnext_mz.utils.tenant_usage.identity"
PROBE_TIMEOUT = 60
HTTP_TIMEOUT = 5
PAID_RECENTLY_DAYS = 60
FUZZY_THRESHOLD = 0.85
STRONG_KEYS = ("site", "nuit", "email")
SITE_COLUMNS = (
	"site", "site_dir", "maintenance_mode", "http_status",
	"company_name", "nuit", "company_email", "company_phone", "responsible", "users",
	"last_login", "invoice_count", "last_invoice_on", "invoice_count_30d", "master_data_count",
	"last_doc_modified", "db_size_mb", "last_backup_on", "probe_error",
	"customer", "contract", "is_signed", "plan", "sub_status", "sub_cancelled_on",
	"invoices", "paid_invoices", "outstanding", "last_billed_on", "last_paid_on", "opportunity", "sales_stage", "lead", "prov_status", "mz_direct",
	"match_keys", "match_quality", "conflicts", "class", "class_reason",
)
CONTROL_COLUMNS = ("record", "name", "company_name", "nuit", "email", "mobile", "contract", "is_signed",
                   "mz_tenant", "sub_status", "opportunity", "sales_stage", "opp_status", "lead")


# ---------------------------------------------------------------- sites on disk

def _sites() -> list[dict]:
	"""Every tenant site directory, live or archived, with what the filesystem knows."""
	bench = get_bench_path()
	found = []
	for state, root in (("live", os.path.join(bench, "sites")), ("archived", os.path.join(bench, "archived", "sites"))):
		if not os.path.isdir(root):
			continue
		for name in sorted(os.listdir(root)):
			path = os.path.join(root, name)
			if not os.path.isdir(path) or not _is_tenant_site(name):
				continue
			found.append({
				"site": name, "site_dir": state, "path": path,
				"maintenance_mode": _maintenance_mode(path),
				"last_backup_on": _last_backup(path),
			})
	return found


def _is_tenant_site(name: str) -> bool:
	# The bare domain itself is a site too (a holding's own instance: erp.kalenyholding.com).
	return any(name.endswith(d) or name == d.lstrip(".") for d in DOMAINS)


def _maintenance_mode(path) -> int:
	try:
		with open(os.path.join(path, "site_config.json")) as f:
			return cint(json.load(f).get("maintenance_mode"))
	except (OSError, ValueError):
		return 0


def _last_backup(path):
	folder = os.path.join(path, "private", "backups")
	try:
		files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".sql.gz")]
	except OSError:
		return None
	if not files:
		return None
	newest = max(files, key=os.path.getmtime)
	return datetime.datetime.fromtimestamp(os.path.getmtime(newest)).replace(microsecond=0)


def _http_status(site: str) -> str:
	"""What the internet sees: 200 / 503 / dns-fail / timeout / conn-fail."""
	import requests

	try:
		r = requests.get(f"https://{site}/api/method/ping", timeout=HTTP_TIMEOUT, allow_redirects=True)
		return str(r.status_code)
	except requests.exceptions.ConnectTimeout:
		return "timeout"
	except requests.exceptions.ConnectionError as exc:
		return "dns-fail" if "Name or service not known" in str(exc) or "nodename" in str(exc) else "conn-fail"
	except requests.exceptions.RequestException as exc:  # pragma: no cover
		return type(exc).__name__


def _probe_identity(site: str) -> dict:
	"""Run erpnext_mz's identity() inside the tenant; never raise."""
	scratch = frappe._dict(log="")
	try:
		raw = run_cmd_capture(
			[get_bench_cmd(), "--site", site, "execute", IDENTITY_METHOD],
			step="identity-probe", prov=scratch, timeout=PROBE_TIMEOUT,
		)
		data = json.loads(raw.strip())
		if isinstance(data, str):
			data = json.loads(data)
		data["error"] = ""
		return data
	except (ProvisioningError, ValueError, TypeError) as exc:
		return {"error": str(exc)[:500], "company": {}, "people": [], "usage": {}}


# ---------------------------------------------------------------- matching

def _norm_name(value) -> str:
	if not value:
		return ""
	s = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().lower()
	s = re.sub(r"\b(lda|limitada|l\.da|sa|s\.a|su|e\.i|ei|unipessoal|sociedade|company|empresa)\b", " ", s)
	return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _norm_mobile(value) -> str:
	digits = re.sub(r"\D", "", str(value or ""))
	return digits[-9:] if len(digits) >= 9 else ""


def _norm_nuit(value) -> str:
	return re.sub(r"\D", "", str(value or ""))


def _norm_email(value) -> str:
	return (value or "").strip().lower()


def _keys_from(site: str, ident: dict) -> dict:
	"""What the site offers to be matched on."""
	company = ident.get("company") or {}
	people = ident.get("people") or []
	return {
		"site": site,
		"slug": site.split(".")[0],
		"nuit": _norm_nuit(company.get("tax_id")),
		"emails": {e for e in (_norm_email(company.get("email")), *(_norm_email(p.get("email")) for p in people)) if e and "@" in e},
		"mobiles": {m for m in (_norm_mobile(company.get("phone_no")), *(_norm_mobile(p.get("mobile_no")) for p in people)) if m},
		"name": _norm_name(company.get("company_name")),
	}


def _match(site: str, ident: dict) -> dict:
	"""Control-side records that point at this site, each with the keys that hit.

	Returns {"customers": {name: {keys}}, "contracts": {...}, "leads": {...},
	"opportunities": {...}, "signups": {...}} — every candidate, never a guess.
	"""
	k = _keys_from(site, ident)
	hits = {t: defaultdict(set) for t in ("customers", "contracts", "leads", "opportunities", "signups")}

	def hit(kind, name, key):
		if name:
			hits[kind][name].add(key)

	# site name
	for c in frappe.get_all("Contract", filters={"docstatus": ("<", 2), "mz_tenant": k["slug"]}, fields=["name", "party_name", "party_type"]):
		hit("contracts", c.name, "site")
		if c.party_type == "Customer":
			hit("customers", c.party_name, "site")
	for c in frappe.get_all("Contract", filters={"docstatus": ("<", 2), "mz_tenant_url": site}, fields=["name", "party_name", "party_type"]):
		hit("contracts", c.name, "site")
		if c.party_type == "Customer":
			hit("customers", c.party_name, "site")
	for p in frappe.get_all("MZ Tenant Provisioning", filters={"site_name": site}, fields=["contract"]):
		hit("contracts", p.contract, "site")
		party = frappe.db.get_value("Contract", p.contract, ["party_type", "party_name"], as_dict=True)
		if party and party.party_type == "Customer":
			hit("customers", party.party_name, "site")
	for s in frappe.get_all("MZ Signup", filters={"subdomain": k["slug"]}, fields=["name", "customer", "lead", "contract", "opportunity"]):
		for kind, name in (("signups", s.name), ("customers", s.customer), ("leads", s.lead),
		                   ("contracts", s.contract), ("opportunities", s.opportunity)):
			hit(kind, name, "site")

	# NUIT
	if k["nuit"]:
		for c in frappe.get_all("Customer", filters={"tax_id": ("like", f"%{k['nuit']}%")}, pluck="name"):
			if _norm_nuit(frappe.db.get_value("Customer", c, "tax_id")) == k["nuit"]:
				hit("customers", c, "nuit")
		for s in frappe.get_all("MZ Signup", filters={"tax_id": ("like", f"%{k['nuit']}%")}, fields=["name", "customer", "lead", "contract", "opportunity"]):
			for kind, name in (("signups", s.name), ("customers", s.customer), ("leads", s.lead),
			                   ("contracts", s.contract), ("opportunities", s.opportunity)):
				hit(kind, name, "nuit")

	# emails
	for email in k["emails"]:
		for c in frappe.get_all("Customer", filters={"email_id": email}, pluck="name"):
			hit("customers", c, "email")
		contacts = frappe.get_all("Contact Email", filters={"email_id": email}, pluck="parent")
		if contacts:
			for link in frappe.get_all("Dynamic Link", filters={"parenttype": "Contact", "parent": ("in", contacts), "link_doctype": "Customer"}, pluck="link_name"):
				hit("customers", link, "email")
		for lead in frappe.get_all("Lead", filters={"email_id": email}, pluck="name"):
			hit("leads", lead, "email")
		for o in frappe.get_all("Opportunity", filters={"contact_email": email}, pluck="name"):
			hit("opportunities", o, "email")
		for s in frappe.get_all("MZ Signup", filters={"email": email}, fields=["name", "customer", "lead", "contract", "opportunity"]):
			for kind, name in (("signups", s.name), ("customers", s.customer), ("leads", s.lead),
			                   ("contracts", s.contract), ("opportunities", s.opportunity)):
				hit(kind, name, "email")
		for c in frappe.get_all("Contract", filters={"docstatus": ("<", 2), "contact_email": email}, fields=["name", "party_name", "party_type"]):
			hit("contracts", c.name, "email")
			if c.party_type == "Customer":
				hit("customers", c.party_name, "email")

	# mobiles (last 9 digits)
	if k["mobiles"]:
		for c in frappe.get_all("Customer", filters={"mobile_no": ("is", "set")}, fields=["name", "mobile_no"]):
			if _norm_mobile(c.mobile_no) in k["mobiles"]:
				hit("customers", c.name, "mobile")
		for p in frappe.get_all("Contact Phone", filters={"phone": ("is", "set")}, fields=["parent", "phone"]):
			if _norm_mobile(p.phone) in k["mobiles"]:
				for link in frappe.get_all("Dynamic Link", filters={"parenttype": "Contact", "parent": p.parent, "link_doctype": "Customer"}, pluck="link_name"):
					hit("customers", link, "mobile")
		for lead in frappe.get_all("Lead", filters={"mobile_no": ("is", "set")}, fields=["name", "mobile_no", "phone"]):
			if _norm_mobile(lead.mobile_no) in k["mobiles"] or _norm_mobile(lead.phone) in k["mobiles"]:
				hit("leads", lead.name, "mobile")

	# company name — exact after normalisation, else fuzzy
	if k["name"]:
		for kind, doctype, field in (("customers", "Customer", "customer_name"), ("leads", "Lead", "company_name"), ("opportunities", "Opportunity", "customer_name")):
			for r in frappe.get_all(doctype, filters={field: ("is", "set")}, fields=["name", field]):
				n = _norm_name(r.get(field))
				if not n:
					continue
				if n == k["name"]:
					hit(kind, r.name, "name")
				elif SequenceMatcher(None, n, k["name"]).ratio() >= FUZZY_THRESHOLD:
					hit(kind, r.name, "fuzzy")

	return {t: {n: sorted(keys) for n, keys in d.items()} for t, d in hits.items()}


def _pick(candidates: dict) -> tuple[str | None, str, str]:
	"""The best candidate, its quality, and a note when the choice was not clean."""
	if not candidates:
		return None, "none", ""

	def strength(item):
		keys = item[1]
		return (sum(k in STRONG_KEYS for k in keys), "name" in keys, "mobile" in keys, len(keys))

	ranked = sorted(candidates.items(), key=strength, reverse=True)
	best, keys = ranked[0]
	quality = "strong" if any(k in STRONG_KEYS for k in keys) else ("name-only" if "name" in keys else ("mobile-only" if "mobile" in keys else "fuzzy"))
	conflicts = ""
	strong_others = [n for n, ks in ranked[1:] if any(k in STRONG_KEYS for k in ks)]
	if strong_others:
		conflicts = "also " + ", ".join(f"{n} ({';'.join(candidates[n])})" for n in strong_others)
	return best, quality, conflicts


# ---------------------------------------------------------------- control side

def _control_side(customer: str | None, contracts_hit: dict, opps_hit: dict, leads_hit: dict) -> dict:
	"""What the control site records about this account."""
	out = {}
	contract = None
	if customer:
		contract = frappe.db.get_value(
			"Contract", {"party_type": "Customer", "party_name": customer, "docstatus": ("<", 2)},
			["name", "is_signed", "mz_subscription_plan", "mz_linked_subscription", "mz_tenant", "docstatus", "mz_direct"],
			as_dict=True, order_by="docstatus desc, creation desc",
		)
	if not customer and contracts_hit:
		# No Customer matched, but a Contract did (by site or email): take the account from it.
		contract = frappe.db.get_value("Contract", sorted(contracts_hit)[-1], ["name", "is_signed", "mz_subscription_plan", "mz_linked_subscription", "mz_tenant", "docstatus", "mz_direct"], as_dict=True)
		party = frappe.db.get_value("Contract", contract.name, ["party_type", "party_name"], as_dict=True) if contract else None
		customer = party.party_name if party and party.party_type == "Customer" else None
	out["customer"] = customer
	out["contract"] = contract.name if contract else None
	out["is_signed"] = cint(contract.is_signed) if contract else None
	out["plan"] = contract.mz_subscription_plan if contract else None
	out["prov_status"] = frappe.db.get_value("MZ Tenant Provisioning", {"contract": contract.name}, "status") if contract else None
	out["mz_direct"] = cint(contract.mz_direct) if contract else None

	sub = None
	if contract and contract.mz_linked_subscription:
		sub = frappe.db.get_value("Subscription", contract.mz_linked_subscription, ["name", "status", "cancelation_date"], as_dict=True)
	if not sub and customer:
		sub = frappe.db.get_value("Subscription", {"party_type": "Customer", "party": customer}, ["name", "status", "cancelation_date"], as_dict=True, order_by="creation desc")
	out["sub_status"] = sub.status if sub else None
	out["sub_cancelled_on"] = sub.cancelation_date if sub else None

	if customer:
		inv = frappe.db.sql(
			"""select count(*) n, sum(outstanding_amount = 0) paid, sum(outstanding_amount) outstanding, max(posting_date) last
			   from `tabSales Invoice` where customer = %s and docstatus = 1 and is_return = 0""",
			customer, as_dict=True,
		)[0]
		out.update(invoices=cint(inv.n), paid_invoices=cint(inv.paid), outstanding=flt(inv.outstanding), last_billed_on=inv.last)
		out["last_paid_on"] = frappe.db.get_value(
			"Payment Entry", {"party_type": "Customer", "party": customer, "docstatus": 1, "payment_type": "Receive"}, "max(posting_date)"
		)
		out["lead"] = frappe.db.get_value("Customer", customer, "lead_name")
	else:
		out.update(invoices=None, paid_invoices=None, outstanding=None, last_billed_on=None, last_paid_on=None, lead=None)
	if not out["lead"] and not customer and leads_hit:
		out["lead"] = sorted(leads_hit)[-1]

	opp = None
	if customer:
		filters = [{"party_name": customer}]
		if out["lead"]:
			filters.append({"opportunity_from": "Lead", "party_name": out["lead"]})
		for f in filters:
			opp = frappe.db.get_value("Opportunity", f, ["name", "sales_stage", "status"], as_dict=True, order_by="creation desc")
			if opp:
				break
	if not opp and not customer and opps_hit:
		opp = frappe.db.get_value("Opportunity", sorted(opps_hit)[-1], ["name", "sales_stage", "status"], as_dict=True)
	out["opportunity"] = opp.name if opp else None
	out["sales_stage"] = f"{opp.sales_stage} ({opp.status})" if opp else None
	return out


# ---------------------------------------------------------------- classification

def _classify(row: dict) -> tuple[str, str]:
	"""A hint, computed from the observed signals only. Returns (class, reason)."""
	if row.get("site_dir") == "archived":
		return "archived_by_hand", "directory under archived/sites"
	if row.get("probe_error"):
		return "unclassified", f"probe failed: {row['probe_error'][:80]}"
	sub = row.get("sub_status")
	outstanding = flt(row.get("outstanding"))
	last_paid = row.get("last_paid_on")
	paid_recently = bool(last_paid) and getdate(last_paid) >= getdate(add_days(nowdate(), -PAID_RECENTLY_DAYS))
	last_login = row.get("last_login")
	active_recently = bool(last_login) and getdate(str(last_login)[:10]) >= getdate(add_days(nowdate(), -30))
	used = cint(row.get("invoice_count")) or cint(row.get("master_data_count")) or last_login
	if sub == "Active" and (outstanding == 0 or paid_recently):
		return "paying", f"subscription Active, outstanding {outstanding:.0f}" + (", paid recently" if paid_recently else "") + ("" if used else " — site never used")
	if outstanding > 0:
		return "debtor", f"subscription {sub or '—'}, outstanding {outstanding:.0f}" + (", still using the site" if active_recently else "")
	if sub in ("Cancelled", "Past Due Date", "Unpaid"):
		return "cancelled_paid_up", f"subscription {sub}, nothing owed" + (", still using the site" if active_recently else "") + ("" if used else ", site never used")
	if not row.get("customer"):
		return "unmatched_site", "no control-site record matched" + (f"; {cint(row.get('invoice_count'))} invoices, last login {str(last_login or '—')[:10]}")
	if not used:
		return "never_used", "no invoices, no own master data, no customer login"
	return "used_unsigned", f"{cint(row.get('invoice_count'))} invoices, last login {str(last_login or '—')[:10]}, no subscription ever"


# ---------------------------------------------------------------- the inventory

def _site_row(s: dict, ident: dict, http: str) -> dict:
	company = ident.get("company") or {}
	people = ident.get("people") or []
	usage = ident.get("usage") or {}
	hits = _match(s["site"], ident)
	customer, quality, conflicts = _pick(hits["customers"])
	control = _control_side(customer, hits["contracts"], hits["opportunities"], hits["leads"])
	if not customer and control.get("customer"):
		quality = "strong"
	keys = set()
	for kind in hits.values():
		for ks in kind.values():
			keys.update(ks)
	responsible = people[0] if people else {}
	row = {
		"site": s["site"], "site_dir": s["site_dir"], "maintenance_mode": s["maintenance_mode"], "http_status": http,
		"company_name": company.get("company_name"), "nuit": company.get("tax_id"),
		"company_email": company.get("email"), "company_phone": company.get("phone_no"),
		"responsible": " · ".join(str(v) for v in (responsible.get("full_name"), responsible.get("email"), responsible.get("mobile_no")) if v) or None,
		"users": usage.get("user_count"), "last_login": usage.get("last_login"),
		"invoice_count": usage.get("invoice_count"), "last_invoice_on": usage.get("last_invoice_date"),
		"invoice_count_30d": usage.get("invoice_count_30d"), "master_data_count": usage.get("master_data_count"),
		"last_doc_modified": usage.get("last_doc_modified"), "db_size_mb": ident.get("db_size_mb"),
		"last_backup_on": s.get("last_backup_on"), "probe_error": ident.get("error") or "",
		**control,
		"match_keys": ";".join(sorted(keys)), "match_quality": quality if control.get("customer") else "none",
		"conflicts": conflicts,
	}
	row["class"], row["class_reason"] = _classify(row)
	return row


def _control_only(matched: dict) -> list[dict]:
	"""Cloud records on the control site that no site claimed."""
	rows = []
	customers = frappe.db.sql(
		"""select distinct c.name, c.customer_name, c.tax_id, c.email_id, c.mobile_no, c.lead_name
		   from `tabCustomer` c
		   where exists (select 1 from `tabContract` k where k.party_type='Customer' and k.party_name=c.name and k.docstatus < 2)
		      or exists (select 1 from `tabSubscription` s where s.party_type='Customer' and s.party=c.name)
		      or exists (select 1 from `tabMZ Signup` g where g.customer=c.name)""",
		as_dict=True,
	)
	for c in customers:
		if c.name in matched["customers"]:
			continue
		k = frappe.db.get_value("Contract", {"party_type": "Customer", "party_name": c.name, "docstatus": ("<", 2)}, ["name", "is_signed", "mz_tenant"], as_dict=True, order_by="creation desc")
		sub = frappe.db.get_value("Subscription", {"party_type": "Customer", "party": c.name}, "status", order_by="creation desc")
		opp = frappe.db.get_value("Opportunity", {"party_name": c.name}, ["name", "sales_stage", "status"], as_dict=True, order_by="creation desc")
		if not opp and c.lead_name:
			opp = frappe.db.get_value("Opportunity", {"opportunity_from": "Lead", "party_name": c.lead_name}, ["name", "sales_stage", "status"], as_dict=True, order_by="creation desc")
		rows.append({
			"record": "Customer", "name": c.name, "company_name": c.customer_name, "nuit": c.tax_id, "email": c.email_id, "mobile": c.mobile_no,
			"contract": k.name if k else None, "is_signed": cint(k.is_signed) if k else None, "mz_tenant": k.mz_tenant if k else None,
			"sub_status": sub, "opportunity": opp.name if opp else None, "sales_stage": opp.sales_stage if opp else None,
			"opp_status": opp.status if opp else None, "lead": c.lead_name,
		})
		if opp:
			matched["opportunities"].add(opp.name)
		if c.lead_name:
			matched["leads"].add(c.lead_name)
	for o in frappe.get_all("Opportunity", filters={"sales_stage": ("like", "Cloud - %")}, fields=["name", "party_name", "opportunity_from", "customer_name", "contact_email", "contact_mobile", "sales_stage", "status"], order_by="creation"):
		if o.name in matched["opportunities"]:
			continue
		rows.append({
			"record": "Opportunity", "name": o.name, "company_name": o.customer_name, "nuit": None, "email": o.contact_email, "mobile": o.contact_mobile,
			"contract": None, "is_signed": None, "mz_tenant": None, "sub_status": None, "opportunity": o.name,
			"sales_stage": o.sales_stage, "opp_status": o.status, "lead": o.party_name if o.opportunity_from == "Lead" else None,
		})
		if o.opportunity_from == "Lead":
			matched["leads"].add(o.party_name)
	for lead in frappe.get_all("Lead", filters={"email_id": ("is", "set")}, fields=["name", "company_name", "email_id", "mobile_no", "status"], order_by="creation"):
		if lead.name in matched["leads"]:
			continue
		if not (frappe.db.exists("Opportunity", {"opportunity_from": "Lead", "party_name": lead.name}) or frappe.db.exists("MZ Signup", {"lead": lead.name}) or frappe.db.exists("Customer", {"lead_name": lead.name})):
			continue
		rows.append({
			"record": "Lead", "name": lead.name, "company_name": lead.company_name, "nuit": None, "email": lead.email_id, "mobile": lead.mobile_no,
			"contract": None, "is_signed": None, "mz_tenant": None, "sub_status": None, "opportunity": None,
			"sales_stage": None, "opp_status": lead.status, "lead": lead.name,
		})
	return rows


def build() -> dict:
	"""The three sheets as lists of dicts. No writes."""
	site_rows, matched = [], {"customers": set(), "opportunities": set(), "leads": set()}
	for s in _sites():
		ident = _probe_identity(s["site"]) if s["site_dir"] == "live" else {"error": "archived: not probed", "company": {}, "people": [], "usage": {}}
		http = _http_status(s["site"])
		row = _site_row(s, ident, http)
		site_rows.append(row)
		if row.get("customer"):
			matched["customers"].add(row["customer"])
		if row.get("opportunity"):
			matched["opportunities"].add(row["opportunity"])
		if row.get("lead"):
			matched["leads"].add(row["lead"])
	control_rows = _control_only(matched)
	summary = defaultdict(int)
	for r in site_rows:
		summary[f"class: {r['class']}"] += 1
		summary[f"match: {r['match_quality']}"] += 1
	for r in control_rows:
		summary[f"control_only: {r['record']}"] += 1
	return {"sites": site_rows, "control_only": control_rows, "summary": [{"key": k, "count": v} for k, v in sorted(summary.items())]}


def to_xlsx(data: dict) -> bytes:
	from openpyxl import Workbook
	from openpyxl.utils import get_column_letter

	wb = Workbook()
	first = True
	for title, columns in (("sites", SITE_COLUMNS), ("control_only", CONTROL_COLUMNS), ("summary", ("key", "count"))):
		ws = wb.active if first else wb.create_sheet()
		first = False
		ws.title = title
		ws.append(list(columns))
		for r in data[title]:
			ws.append([_cell(r.get(c)) for c in columns])
		ws.freeze_panes = "B2"
		for i, c in enumerate(columns, 1):
			ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(c) + 4))
	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


def _cell(v):
	if v is None or isinstance(v, (int, float, str)):
		return v
	return str(v)


def inventory(attach: int = 1) -> dict:
	"""Build the inventory, attach it to MZ SaaS Settings, print the URL and the counts."""
	data = build()
	name = f"tenant_inventory_{nowdate().replace('-', '')}.xlsx"
	url = None
	if cint(attach):
		for old in frappe.get_all("File", filters={"file_name": name, "attached_to_doctype": "MZ SaaS Settings"}, pluck="name"):
			frappe.delete_doc("File", old, ignore_permissions=True, force=True)
		f = frappe.get_doc({
			"doctype": "File", "file_name": name, "is_private": 1,
			"attached_to_doctype": "MZ SaaS Settings", "attached_to_name": "MZ SaaS Settings",
			"content": to_xlsx(data),
		}).insert(ignore_permissions=True)
		frappe.db.commit()
		url = frappe.utils.get_url(f.file_url)
	print("\n".join(f"{r['key']:<32}{r['count']:>5}" for r in data["summary"]))
	if url:
		print(f"\n{url}")
	return {"file_url": url, "summary": data["summary"], "sites": len(data["sites"]), "control_only": len(data["control_only"])}


# ---------------------------------------------------------------- phase 2: acting

def _contract_list(contracts) -> list[str]:
	"""A comma list, a Python list, or "@file.csv" (column `contract`)."""
	if isinstance(contracts, (list, tuple)):
		return [c.strip() for c in contracts if c and str(c).strip()]
	contracts = str(contracts or "").strip()
	if contracts.startswith("@"):
		import csv

		with open(contracts[1:], newline="") as f:
			return sorted({r["contract"].strip() for r in csv.DictReader(f) if (r.get("contract") or "").strip()})
	return [c.strip() for c in contracts.split(",") if c.strip()]


def _live_site_for(doc) -> str | None:
	"""The live site directory this contract points at — from its own fields only."""
	candidates = [doc.mz_tenant_url] if doc.mz_tenant_url else []
	if doc.mz_tenant:
		from ai_saas.saas.provisioning import domain_for

		candidates.append(doc.mz_tenant + domain_for(doc.mz_domain))
	live = {s["site"] for s in _sites() if s["site_dir"] == "live"}
	for site in candidates:
		if site in live:
			return site
	return None


def _register_provisioning(contract_doc, site: str) -> str:
	"""The row that makes the account visible to the lifecycle. Inserting one runs no
	command and sends nothing (the doctype has no hooks; only Queued-ish rows are retried)."""
	existing = frappe.db.get_value("MZ Tenant Provisioning", {"contract": contract_doc.name}, "name")
	if existing:
		return existing
	row = frappe.get_doc({
		"doctype": "MZ Tenant Provisioning", "contract": contract_doc.name,
		"tenant_slug": contract_doc.mz_tenant or site.split(".")[0], "site_name": site,
		"status": "Active", "customer_name": contract_doc.party_name,
		"contact_email": contract_doc.contact_email,
		"provisioned_at": contract_doc.creation,
		"log": f"{nowdate()} registado pela migração do funil antigo (legacy_migration)",
	})
	row.insert(ignore_permissions=True)
	return row.name


def _report_activated(customer: str, contract_name: str) -> str:
	"""The ledger entry for a signed legacy account: existing Opportunity moved to
	Activated/Converted, or one created already there so the ledger is complete."""
	from ai_saas.saas import crm

	opportunity = crm.find_opportunity(contract_name)
	if not opportunity:
		from ai_saas.saas.contract_lifecycle import _get_company

		opp = frappe.get_doc({
			"doctype": "Opportunity", "opportunity_from": "Customer", "party_name": customer,
			"company": _get_company(), "transaction_date": nowdate(),
			"sales_stage": crm.STAGE_ACTIVATED,
		})
		opp.insert(ignore_permissions=True)
		opportunity = opp.name
	crm.report(opportunity, crm.STAGE_ACTIVATED, status="Converted")
	return opportunity


def activate(contracts, dry_run: int = 1):
	"""Bring signed legacy accounts into the lifecycle, reusing every existing record.

	Per contract: provisioning row (status Active), Contract linked to its existing
	Subscription (mz_linked_subscription, mz_billing_start, plan/tenant fields filled
	where empty, mz_direct = 1), Customer primaries + commercial group, Opportunity at
	Activated/Converted. All writes via db_set — the signature hooks must NOT run
	(they would create a second Subscription and invoice today). No customer email.

		bench --site <site> execute ai_saas.saas.legacy_migration.activate \
			--kwargs "{'contracts': '@sheet.csv', 'dry_run': 1}"
	"""
	from ai_saas.saas.alerts import notify_ops
	from ai_saas.saas.contract_lifecycle import _move_customer_to_commercial_group
	from ai_saas.saas.party import ensure_customer_primaries
	from ai_saas.saas.tenant_lifecycle import _attempt

	dry_run = cint(dry_run)
	lines = []

	def plan_for(name):
		doc = frappe.get_doc("Contract", name)
		if doc.docstatus != 1 or not doc.is_signed or doc.party_type != "Customer":
			return None, f"{name}: IGNORADO — precisa de contrato submetido e assinado de um Customer."
		site = _live_site_for(doc)
		if not site:
			return None, f"{name}: IGNORADO — nenhum site vivo encontrado (mz_tenant/mz_tenant_url)."
		subs = frappe.get_all(
			"Subscription",
			filters={"party_type": "Customer", "party": doc.party_name, "status": ("!=", "Cancelled")},
			fields=["name", "start_date"], order_by="creation desc",
		)
		if len(subs) > 1:
			return None, f"{name}: IGNORADO — {len(subs)} subscrições activas para {doc.party_name}; decidir à mão."
		return (doc, site, subs[0] if subs else None), None

	def act(doc, site, sub):
		prov = _register_provisioning(doc, site)
		updates = {"mz_direct": 1}
		if sub:
			updates["mz_linked_subscription"] = sub.name
			if not doc.mz_billing_start:
				updates["mz_billing_start"] = sub.start_date
			if not doc.mz_subscription_plan:
				plan = frappe.db.get_value("Subscription Plan Detail", {"parent": sub.name}, "plan")
				if plan:
					updates["mz_subscription_plan"] = plan
		if not doc.mz_tenant_url:
			updates["mz_tenant_url"] = site
		if not doc.mz_tenant:
			updates["mz_tenant"] = site.split(".")[0]
		doc.db_set(updates, update_modified=False)
		ensure_customer_primaries(doc.party_name)
		_move_customer_to_commercial_group(frappe._dict(party_name=doc.party_name))
		opportunity = _report_activated(doc.party_name, doc.name)
		lines.append(
			f"{doc.name}: {site} → prov {prov}, subscrição {sub.name if sub else '— (nenhuma ligada)'}, "
			f"oportunidade {opportunity}"
		)

	for name in _contract_list(contracts):
		planned, problem = plan_for(name)
		if problem:
			lines.append(problem)
			continue
		doc, site, sub = planned
		if dry_run:
			lines.append(
				f"{name}: [dry-run] site {site}, subscrição {sub.name if sub else '— (nenhuma: não liga)'}, "
				f"prov {'existe' if frappe.db.exists('MZ Tenant Provisioning', {'contract': name}) else 'a criar'}"
			)
			continue
		_attempt(lines, name, lambda doc=doc, site=site, sub=sub: act(doc, site, sub))

	digest = "\n".join(lines) or "nada a fazer"
	print(digest)
	if not dry_run and lines:
		notify_ops("Migração de contas do funil antigo", "<br>".join(frappe.utils.escape_html(x) for x in lines))
	return lines


def create_account(customer_name, site, plan=None, start_date=None, email=None, dry_run: int = 1):
	"""A direct account whose site already exists (a holding's own instance, a yearly
	deal paid outside): register the provisioning row FIRST — so the submit hook's
	provision_tenant no-ops — then create/reuse Customer and Contact and submit a
	Contract already signed. With `plan`, the normal hook creates the Subscription
	(billing starts at max(start_date, today) — never back-dated); without it, the
	engine-silent CCM/partner shape. The delivery email is never sent (nothing provisions).
	"""
	from ai_saas.saas import crm
	from ai_saas.saas.contract_lifecycle import _get_company

	dry_run = cint(dry_run)
	live = {s["site"] for s in _sites() if s["site_dir"] == "live"}
	if site not in live:
		frappe.throw(f"O site {site} não existe vivo neste bench.")
	if plan and not frappe.db.exists("Subscription Plan", plan):
		frappe.throw(f"Plano desconhecido: {plan}")
	taken = frappe.db.get_value("MZ Tenant Provisioning", {"site_name": site}, "contract")
	if taken:
		frappe.throw(f"O site {site} já pertence ao contrato {taken} — use activate().")

	ident = _probe_identity(site)
	responsible = (ident.get("people") or [{}])[0]
	contact_email = email or responsible.get("email") or (ident.get("company") or {}).get("email")
	if dry_run:
		line = (f"[dry-run] {customer_name} @ {site}: customer "
		        f"{'existe' if frappe.db.exists('Customer', {'customer_name': customer_name}) else 'a criar'}, "
		        f"contacto {contact_email or '—'}, plano {plan or '— (sem subscrição)'}, início {start_date or nowdate()}")
		print(line)
		return line

	customer = frappe.db.get_value("Customer", {"customer_name": customer_name}, "name")
	if not customer:
		customer = frappe.get_doc({
			"doctype": "Customer", "customer_name": customer_name, "customer_type": "Company",
			"customer_group": frappe.db.get_single_value("Selling Settings", "customer_group")
			or frappe.db.get_value("Customer Group", {"is_group": 0}, "name"),
			"territory": frappe.db.get_single_value("Selling Settings", "territory")
			or frappe.db.get_value("Territory", {"is_group": 0}, "name"),
			"tax_id": _norm_nuit((ident.get("company") or {}).get("tax_id")) or None,
		}).insert(ignore_permissions=True).name
	if contact_email and not frappe.db.get_value("Customer", customer, "customer_primary_contact"):
		from ai_saas.saas.party import set_customer_primaries

		contact = frappe.get_doc({
			"doctype": "Contact", "first_name": responsible.get("full_name") or customer_name,
			"is_primary_contact": 1,
			"email_ids": [{"email_id": contact_email, "is_primary": 1}],
			"phone_nos": [{"phone": responsible.get("mobile_no"), "is_primary_mobile_no": 1}]
			if responsible.get("mobile_no") else [],
			"links": [{"link_doctype": "Customer", "link_name": customer}],
		}).insert(ignore_permissions=True)
		set_customer_primaries(customer, contact=contact.name, email=contact_email,
		                       mobile=responsible.get("mobile_no"))

	contract = frappe.get_doc({
		"doctype": "Contract", "party_type": "Customer", "party_name": customer,
		"start_date": start_date or nowdate(), "is_signed": 1, "signed_on": frappe.utils.now_datetime(),
		"mz_direct": 1, "mz_subscription_plan": plan, "mz_tenant": site.split(".")[0],
		"mz_tenant_url": site,
		"contract_terms": "Contrato de venda directa — negociado e assinado fora do sistema.",
	})
	contract.insert(ignore_permissions=True)
	_register_provisioning(contract, site)
	if not crm.find_opportunity(contract.name):
		frappe.get_doc({
			"doctype": "Opportunity", "opportunity_from": "Customer", "party_name": customer,
			"company": _get_company(), "transaction_date": nowdate(),
			"sales_stage": crm.STAGE_ACCOUNT_CREATED,
		}).insert(ignore_permissions=True)
	contract.submit()
	if not plan:
		# on_contract_submitted returns before the ledger and the group move when there is
		# no plan (nothing to bill, nothing to provision) — a plan-less direct account
		# (holding/partner) is still a converted customer, so record it here.
		from ai_saas.saas.contract_lifecycle import _move_customer_to_commercial_group

		_move_customer_to_commercial_group(frappe._dict(party_name=customer))
		_report_activated(customer, contract.name)
	frappe.db.commit()
	result = {"customer": customer, "contract": contract.name,
	          "subscription": frappe.db.get_value("Contract", contract.name, "mz_linked_subscription")}
	print(result)
	return result
